from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from daily_movers.models import ReportRow


_HEADERS = [
    "Ticker",
    "Name",
    "Market",
    "Open",
    "Close",
    "Price",
    "Abs Change",
    "% Change",
    "Volume",
    "Currency",
    "Exchange",
    "Sector",
    "Industry",
    "Earnings Date",
    "Action",
    "Confidence",
    "Sentiment",
    "Needs Review",
    "Needs Review Reason",
    "Recommendation Tags",
    "Why It Moved",
    "Top Headline",
    "Headline URL",
    "Trend (ASCII Sparkline)",
    "Decision Trace",
    "Rules Triggered",
    "Evidence Titles",
    "Numeric Signals",
    "Provenance URLs",
    "Model Used",
    "Errors",
]


def write_excel_report(*, rows: list[ReportRow], out_path: Path) -> None:
    """Write an Excel report for a run.

    - One row per ticker, with flattened enrichment + analysis fields.
    - Adds hyperlinks to Yahoo quote pages and (when safe) to the top headline.
    - Includes a Highlights sheet with "Top Pick" and "Most Potential".
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Movers"
    ws.append(_HEADERS)

    header_fill = PatternFill(start_color="1D3557", end_color="1D3557", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx in range(1, len(_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, report_row in enumerate(rows, start=2):
        flat = report_row.to_flat_dict()
        quote_url = f"https://finance.yahoo.com/quote/{flat['ticker']}"
        top_headline_url = flat.get("headline_url")
        trend_points = flat.get("trend_points") or []
        trend_ascii = _ascii_sparkline(trend_points)
        market = _detect_market_label(flat.get("ticker", ""), report_row.ticker.market)

        values = [
            flat.get("ticker"),
            flat.get("name"),
            market,
            flat.get("open_price"),
            flat.get("close_price"),
            flat.get("price"),
            flat.get("abs_change"),
            flat.get("pct_change"),
            flat.get("volume"),
            flat.get("currency"),
            flat.get("exchange"),
            flat.get("sector"),
            flat.get("industry"),
            flat.get("earnings_date"),
            flat.get("action"),
            flat.get("confidence"),
            flat.get("sentiment"),
            "YES" if flat.get("needs_review") else "NO",
            flat.get("needs_review_reason"),
            flat.get("recommendation_tags"),
            flat.get("why_it_moved"),
            flat.get("top_headline"),
            top_headline_url,
            trend_ascii,
            flat.get("decision_trace"),
            flat.get("rules_triggered"),
            flat.get("evidence_titles"),
            flat.get("numeric_signals"),
            flat.get("provenance_urls"),
            report_row.analysis.model_used,
            flat.get("errors"),
        ]

        ws.append(values)

        ticker_col = _HEADERS.index("Ticker") + 1
        ticker_cell = ws.cell(row=row_idx, column=ticker_col)
        ticker_cell.hyperlink = quote_url
        ticker_cell.style = "Hyperlink"

        headline_url_col = _HEADERS.index("Headline URL") + 1
        headline_url_cell = ws.cell(row=row_idx, column=headline_url_col)
        if top_headline_url and _is_safe_url(str(top_headline_url)):
            headline_url_cell.hyperlink = str(top_headline_url)
            headline_url_cell.style = "Hyperlink"

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{max(2, len(rows) + 1)}"

    # % Change column conditional formatting (column F now).
    pct_col = get_column_letter(_HEADERS.index("% Change") + 1)
    ws.conditional_formatting.add(
        f"{pct_col}2:{pct_col}{max(2, len(rows) + 1)}",
        ColorScaleRule(
            start_type="num",
            start_value=-10,
            start_color="F94144",
            mid_type="num",
            mid_value=0,
            mid_color="F9C74F",
            end_type="num",
            end_value=10,
            end_color="43AA8B",
        ),
    )

    # Confidence column conditional formatting (column N now).
    conf_col = get_column_letter(_HEADERS.index("Confidence") + 1)
    ws.conditional_formatting.add(
        f"{conf_col}2:{conf_col}{max(2, len(rows) + 1)}",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="F94144",
            mid_type="num",
            mid_value=0.75,
            mid_color="F9C74F",
            end_type="num",
            end_value=1,
            end_color="43AA8B",
        ),
    )

    widths_by_header = {
        "Ticker": 12,
        "Name": 28,
        "Market": 10,
        "Open": 12,
        "Close": 12,
        "Price": 12,
        "Abs Change": 12,
        "% Change": 12,
        "Volume": 14,
        "Currency": 10,
        "Exchange": 12,
        "Sector": 18,
        "Industry": 22,
        "Earnings Date": 14,
        "Action": 10,
        "Confidence": 11,
        "Sentiment": 10,
        "Needs Review": 13,
        "Needs Review Reason": 26,
        "Recommendation Tags": 28,
        "Why It Moved": 48,
        "Top Headline": 40,
        "Headline URL": 40,
        "Trend (ASCII Sparkline)": 22,
        "Decision Trace": 48,
        "Rules Triggered": 32,
        "Evidence Titles": 48,
        "Numeric Signals": 48,
        "Provenance URLs": 48,
        "Model Used": 18,
        "Errors": 48,
    }
    for idx, header in enumerate(_HEADERS, start=1):
        width = widths_by_header.get(header)
        if width:
            ws.column_dimensions[get_column_letter(idx)].width = width

    # --- Highlights sheet ---
    hl_ws = wb.create_sheet(title="Highlights")
    hl_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    hl_font_white = Font(color="FFFFFF", bold=True, size=12)
    hl_font_gold = Font(color="FFD700", bold=True, size=13)
    hl_font_blue = Font(color="1565C0", bold=True, size=13)

    hl_ws.merge_cells("A1:D1")
    hl_ws["A1"] = "🏆 TOP RECOMMENDED PICK"
    hl_ws["A1"].fill = hl_fill
    hl_ws["A1"].font = hl_font_gold

    top_pick = None
    most_potential = None
    for r in rows:
        if "top_pick_candidate" in r.recommendation_tags:
            if top_pick is None or r.analysis.confidence > top_pick.analysis.confidence:
                top_pick = r
        if "most_potential_candidate" in r.recommendation_tags:
            if most_potential is None or r.analysis.sentiment > most_potential.analysis.sentiment:
                most_potential = r

    if top_pick:
        hl_ws["A2"] = "Ticker"
        hl_ws["B2"] = top_pick.ticker.ticker
        hl_ws["A3"] = "Company"
        hl_ws["B3"] = top_pick.ticker.name
        hl_ws["A4"] = "Action"
        hl_ws["B4"] = top_pick.analysis.action.value
        hl_ws["A5"] = "Confidence"
        hl_ws["B5"] = top_pick.analysis.confidence
        hl_ws["A6"] = "% Change"
        hl_ws["B6"] = top_pick.ticker.pct_change
        hl_ws["A7"] = "Why"
        hl_ws["B7"] = top_pick.analysis.why_it_moved
    else:
        hl_ws["A2"] = "No strong BUY signal with high confidence found."

    hl_ws.merge_cells("A9:D9")
    hl_ws["A9"] = "📈 MOST POTENTIAL"
    hl_ws["A9"].fill = PatternFill(start_color="0277BD", end_color="0277BD", fill_type="solid")
    hl_ws["A9"].font = hl_font_white

    if most_potential:
        hl_ws["A10"] = "Ticker"
        hl_ws["B10"] = most_potential.ticker.ticker
        hl_ws["A11"] = "Company"
        hl_ws["B11"] = most_potential.ticker.name
        hl_ws["A12"] = "Action"
        hl_ws["B12"] = most_potential.analysis.action.value
        hl_ws["A13"] = "Sentiment"
        hl_ws["B13"] = most_potential.analysis.sentiment
        hl_ws["A14"] = "% Change"
        hl_ws["B14"] = most_potential.ticker.pct_change
        hl_ws["A15"] = "Why"
        hl_ws["B15"] = most_potential.analysis.why_it_moved
    else:
        hl_ws["A10"] = "No moderate-confidence upside candidate identified."

    # Market Breakdown section
    hl_ws.merge_cells("A17:D17")
    hl_ws["A17"] = "🌍 MARKET BREAKDOWN"
    hl_ws["A17"].fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    hl_ws["A17"].font = hl_font_white

    market_counts: dict[str, int] = {}
    for r in rows:
        mkt = _detect_market_label(r.ticker.ticker, r.ticker.market)
        market_counts[mkt] = market_counts.get(mkt, 0) + 1

    hl_row = 18
    for mkt, count in sorted(market_counts.items(), key=lambda x: -x[1]):
        hl_ws[f"A{hl_row}"] = mkt.upper()
        hl_ws[f"B{hl_row}"] = count
        hl_row += 1

    hl_ws.column_dimensions["A"].width = 20
    hl_ws.column_dimensions["B"].width = 60

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _is_safe_url(value: str) -> bool:
    lowered = value.strip().lower()
    return lowered.startswith("http://") or lowered.startswith("https://")


def _ascii_sparkline(values: list[float]) -> str:
    if not values:
        return "sparkline unavailable"

    min_v = min(values)
    max_v = max(values)
    if min_v == max_v:
        return "=" * min(12, len(values))

    palette = [".", ":", "-", "=", "+", "*", "#"]
    out: list[str] = []
    for value in values[-12:]:
        ratio = (value - min_v) / (max_v - min_v)
        idx = int(ratio * (len(palette) - 1))
        out.append(palette[idx])
    return "".join(out)


def _detect_market_label(ticker: str, market_hint: str | None) -> str:
    """Detect market from ticker suffix and market hint."""
    t = ticker.upper()
    if t.endswith(".TA"):
        return "TASE"
    if t.endswith(".L"):
        return "UK"
    if any(t.endswith(s) for s in (".PA", ".DE", ".AS", ".MI", ".MC")):
        return "EU"
    if "-USD" in t or t in ("BTC", "ETH", "SOL", "XRP", "DOGE", "BNB"):
        return "Crypto"
    if market_hint:
        hint = market_hint.lower()
        if hint in ("il", "tase"):
            return "TASE"
        if hint == "uk":
            return "UK"
        if hint == "eu":
            return "EU"
        if hint == "crypto":
            return "Crypto"
    return "US"
